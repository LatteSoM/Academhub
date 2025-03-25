import os
import re
import sys
from collections import defaultdict
from pathlib import Path

from django.db.models import Prefetch, Q
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from unidecode import unidecode

from Academhub import settings
from Academhub.modules.documentGenPars import GradebookDocumentGenerator
from Gradebook.forms import *
from Gradebook.forms import GenerateGradebookForm, GetStatisticksGradebookForm
from Gradebook.tables import *
from Gradebook.filters import *
from django.utils import timezone
from django.contrib import messages
from django.http import HttpResponse
from Gradebook.mixins import GradeBookMixin
from Gradebook.tables import TeacherGradeBookTable
from Academhub.utils import getpermission, getpattern
from Gradebook.filters import GradeBookTeachersFilter
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import permission_required
from Academhub.models import GradebookStudents, Gradebook, CustomUser, SubTable, Button, Discipline, Student, \
    GroupStudents, ClockCell, Curriculum
from Academhub.generic import BulkUpdateView, ObjectTableView, ObjectDetailView, ObjectUpdateView, ObjectCreateView, \
    ObjectTemplateView


#
# Create your views here.
#

__all__ = (
    'download_report',
    'GradebookTableView',
    'GradebookClosedList',
    'GradebookDetailView',
    'GradebookCreateView',
    'GradebookUpdateView',
    'check_and_open_gradebook',
    'TeachersGradeBookTableView',
    'GradebookStudentBulkUpdateView',
)

class GradebookStudentBulkUpdateView(BulkUpdateView):
    model = GradebookStudents
    form_class = GradebookStudentsForm
    template_name = 'Gradebook/create/grade_book_students.html'

    def dispatch(self, request, *args, **kwargs):
        self.gradebook_pk = self.kwargs.get('pk', None)
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        return self.model.objects.filter(gradebook__pk = self.gradebook_pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        self.gradebook = get_object_or_404(Gradebook, pk=self.gradebook_pk)
        context['gradebook'] = self.gradebook
        return context
    
    def save_form(self, request):
        form = super().save_form(request)

        if form.is_valid():
            gradebook = get_object_or_404(
                Gradebook, 
                pk=self.gradebook_pk
            )
            gradebook.status = Gradebook.STATUS_CHOICE[2][1]
            gradebook.save()

        return form

    def post(self, request, *args, **kwargs):
        formset = self.save_form(request)

        if formset.is_valid():
            user = get_object_or_404(CustomUser, pk=self.request.user.pk)

            if user.is_teacher:
                response = redirect('gradebookteacher_list')
            else:
                response = redirect('gradebook_list')
            response['Location'] += '?is_filled=true'
            return response
        return super().post(request, *args, **kwargs)


class GradebookGenerateView(ObjectTemplateView):
    """Отдельный класс для обработки генерации"""
    template_name = 'Gradebook/inc/generate_form.html'

    def get(self, request):
        form = GenerateGradebookForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = GenerateGradebookForm(request.POST)  # Создаем экземпляр формы

        if form.is_valid():  # Проверяем валидность ФОРМЫ, а не представления
            try:
                semester = form.cleaned_data['semester']  # Данные берем из формы
                result = perform_generation(semester)

                if result['generated'] > 0 and result['errors'] == 0:
                    messages.success(request,
                                     f"Успешно сгенерировано {result['generated']} ведомостей для {semester} семестра!")
                elif result['generated'] > 0:
                    messages.warning(request,
                                     f"Сгенерировано {result['generated']} ведомостей для {semester} семестра. "
                                     f"Ошибок: {result['errors']}")
                else:
                    messages.error(request,
                                   "Не удалось сгенерировать ни одной ведомости!")

                return redirect('gradebook_list')

            except Exception as e:
                messages.error(request, f"Ошибка генерации: {str(e)}")
                return redirect('gradebook_list')

        # Если форма невалидна, показываем ее с ошибками
        return render(request, self.template_name, {'form': form})


def perform_generation(semester_number):
    generated = 1
    errors = 0
    # TODO: генерация ведомостей по уч плану
    return {'generated': generated, 'errors': errors}


def fetch_group_semester_attestation_table(semester, group):
    """
    Функция для передачи данных на страницу статистики
    """

    # 1. Получаем всех студентов группы
    students = Student.objects.filter(group=group).order_by('full_name')

    # 2. Получаем все ведомости для группы и семестра
    gradebooks = Gradebook.objects.filter(
        group=group,
        semester_number=semester
    ).prefetch_related(
        Prefetch('gradebookstudents_set',
                 queryset=GradebookStudents.objects.select_related('student'),
                 to_attr='grades_list'),
        'discipline'
    )


    curriculum = Curriculum.objects.filter(qualification_name=group.qualification.name, admission_year=group.year_create)

    # course = (int(semester) + 1) // 2

    curriculum_items = ClockCell.objects.filter(
        curriculum=curriculum.first(),
        module=None,
        term=int(semester),
    ).filter(
        Q(
            code_of_type_work__in=[
                'Экзамен',
                'Другие формы контроля',
                'Зачет с оценкой',
                'Курсовая работа'
            ]
        ) |
        Q(
            discipline__code__contains='ГИА',
            discipline__name='Проведение государственных экзаменов'
        ) |
        Q(discipline__code__contains='ПДП')
    )

    disciplines = []
    for item in curriculum_items:

        if int(item.term) == int(semester) and item.discipline not in disciplines:
            disciplines.append(Discipline.objects.get(pk=item.discipline_id))
            print(Discipline.objects.get(pk=item.discipline_id))

    # Создаем матрицу студентов с ключами для всех дисциплин
    grade_matrix = {}
    for student in Student.objects.filter(group=group):
        grade_matrix[student.id] = {
            discipline.id: "-" for discipline in disciplines  # Инициализируем все дисциплины как None
        }

    # Проходим по всем ведомостям (gradebooks) и заполняем оценки
    for gb in gradebooks:
        discipline_id = gb.discipline.id

        # Фильтруем ведомости по текущему семестру и группе
        if gb.semester_number != int(semester) or gb.group != group:
            continue

        # Получаем все оценки из этой ведомости
        grade_entries = GradebookStudents.objects.filter(gradebook=gb)

        # Обновляем матрицу оценок
        for grade_entry in grade_entries:
            student_id = grade_entry.student.id
            if student_id in grade_matrix:
                # Заполняем оценку только если ведомость закрыта
                if gb.status == "Закрыта":
                    grade_matrix[student_id][discipline_id] = grade_entry

        group = gb.group


    group = GroupStudents.objects.filter(full_name=group)



    return {
        'students': students,
        'disciplines': disciplines,
        'grades': grade_matrix,
        'semestr': semester,
        'group': group,
        'gradebooks': gradebooks,
        'curriculum_items': curriculum_items,
    }



class ViewStudentsGradesForSemester(ObjectTemplateView):
    """Отдельный класс для обработки статистической таблицы студентов"""
    template_name = 'Gradebook/inc/statisticks_form.html'

    def get(self, request):
        form = GetStatisticksGradebookForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = GetStatisticksGradebookForm(request.POST)  # Создаем экземпляр формы

        if form.is_valid():  # Проверяем валидность ФОРМЫ
            try:
                semester = form.cleaned_data['semester']  # Данные берем из формы
                group = form.cleaned_data['group']

                data = fetch_group_semester_attestation_table(semester, group)

                return render(request, 'Gradebook/list/group_semester_grades.html', {
                    'students': data['students'],
                    'disciplines': data['disciplines'],
                    'grades': data['grades'],
                    'semestr': data['semestr'],
                    'group': data['group'],
                    'curriculum_items': data['curriculum_items'],
                })

            except Exception as e:
                print(e)
                messages.error(request,
                               "Не удалось сгенерировать статистику!")

        # Если форма невалидна, показываем ее с ошибками
        return render(request, self.template_name, {'form': form})


def export_grades(request):
    try:
        # Получение и валидация параметров
        semester = request.GET.get('semester')
        group_id = request.GET.get('group')

        if not semester or not group_id:
            raise ValueError("Не указаны обязательные параметры")

        group = get_object_or_404(GroupStudents, id=int(group_id))
        semester_number = int(semester)

    except (ValueError, TypeError) as e:
        return HttpResponse(f"Ошибка параметров: {str(e)}", status=400)
    except Exception as e:
        return HttpResponse(f"Ошибка сервера: {str(e)}", status=500)

    # Получение данных
    data = fetch_group_semester_attestation_table(semester_number, group)

    # Создание Excel-документа
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводная ведомость"

    # Стили оформления
    thin_side = Side(style='thin')
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='D9E1F2', fill_type='solid')
    header_font = Font(bold=True)

    # Цвета заливки
    FILLS = {
        'Первичная сдача': PatternFill(start_color='D4EDDA', fill_type='solid'),
        'Пересдача': PatternFill(start_color='CCE5FF', fill_type='solid'),
        'Комиссия': PatternFill(start_color='E2D9F3', fill_type='solid'),
        'неуд': PatternFill(start_color='FFC7CE', fill_type='solid')
    }

    # Заголовок документа
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f"Сводная ведомость успеваемости студентов\nГруппа {group.full_name}\nСеместр {semester}"
    title_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    title_cell.font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 60


    ws.cell(row=5, column=1, value="№ п/п").font = header_font
    ws.cell(row=5, column=2, value="Ф.И.О. студента").font = header_font
    ws.cell(row=5, column=3, value="Основа обучения").font = header_font

    # Группировка дисциплин по типам аттестации из Curriculum
    TYPE_MAPPING = {
        'Экзамен': 'Экзамены',
        'Зачет с оценкой': 'Диф. зачёты',
        'Курсовая работа': 'Курсовые работы',
        'Другие формы контроля': 'Другие формы'
    }

    course = (semester_number + 1) // 2
    current_term = 1 if semester_number % 2 == 1 else 2

    discipline_groups = defaultdict(list)
    for curriculum_item in data['curriculum_items']:

        if curriculum_item.term != current_term or curriculum_item.course != course:
            continue

        discipline = curriculum_item.discipline
        category = TYPE_MAPPING.get(
            curriculum_item.code_of_type_work,
            'Другие формы'
        )

        if discipline.id not in {d.id for d in discipline_groups[category]}:
            discipline_groups[category].append(discipline)

    # Формирование структуры таблицы
    headers = ['№ п/п', 'Ф.И.О. студента', 'Основа обучения']
    column_map = {}
    current_col = 4

    # Порядок отображения категорий
    categories_order = [
        'Экзамены',
        'Диф. зачёты',
        'Курсовые работы',
        'Другие формы'
    ]

    for category in categories_order:
        disciplines = sorted(
            discipline_groups.get(category, []),
            key=lambda x: x.name
        )

        if not disciplines:
            continue

        # Объединение ячеек для категории
        end_col = current_col + len(disciplines) - 1
        ws.merge_cells(
            start_row=4,
            start_column=current_col,
            end_row=4,
            end_column=end_col
        )
        cell = ws.cell(row=4, column=current_col, value=category)
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

        # Добавление дисциплин
        for idx, disc in enumerate(disciplines, start=current_col):
            cell = ws.cell(row=5, column=idx, value=disc.name)
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            column_map[disc.id] = idx

        current_col += len(disciplines)

    # Заполнение данных студентов
    for row_idx, student in enumerate(data['students'], start=6):
        # Базовые данные
        ws.cell(row=row_idx, column=1, value=row_idx - 5).alignment = center_alignment
        ws.cell(row=row_idx, column=2, value=student.full_name)
        ws.cell(row=row_idx, column=3, value=student.education_basis).alignment = center_alignment

        # Оценки
        for category in categories_order:
            for disc in discipline_groups.get(category, []):
                col = column_map.get(disc.id)
                if not col:
                    continue

                grade_entry = data['grades'].get(student.id, {}).get(disc.id)

                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = center_alignment
                cell.border = thin_border

                if grade_entry:
                    if grade_entry == "-":
                        cell.value = grade_entry
                    else:
                        cell.value = grade_entry.grade

                        if str(grade_entry.grade).lower() in ['неудовлетворительно', 'неявка', 'н/я', '2']:
                            cell.fill = FILLS['неуд']
                        elif grade_entry.gradebook.type_of_grade_book in FILLS:
                            cell.fill = FILLS[grade_entry.gradebook.type_of_grade_book]

    # Настройка ширины столбцов
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20 if col_idx > 3 else 15

    # Применение стилей ко всем ячейкам
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row > 5 and cell.column in [1, 3]:
                cell.alignment = center_alignment

    # Формирование имени файла
    group_name_clean = unidecode(group.full_name)
    group_name_clean = re.sub(r'\W+', '_', group_name_clean)[:50]
    filename = f"Vedomost_{group_name_clean}_sem{semester}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

class GradebookTableView(ObjectTableView):
    """
    Класс для отображения таблицы учебных журналов.
    """
    table_class = GradebookTable
    filterset_class = GradebookFilter
    queryset = Gradebook.objects.all()

    buttons = [
        Button (
            id='add',
            name = 'Добавить',
            link_name = getpattern(Gradebook, 'add'),
            permission = getpermission(Gradebook, 'add'),
        ),
        Button(
            id = 'generate',
            name = 'Сгенерировать ведомости',
            link_name = 'gradebook_generate',
            permission = getpermission(Gradebook, 'add'),
        )
    ]
    queryset = Gradebook.objects.exclude(status=Gradebook.STATUS_CHOICE[3][1])

    def get_table_class(self):
        if self.request.GET.get("mobile") == "1":
            return GradebookMobileTable
        return GradebookTable

    def get(self, request, *args, **kwargs):

        is_closed = request.GET.get('is_closed', None)
        is_filled = request.GET.get('is_filled', None)

        if is_closed:
            messages.success(request, "Ведомость успешно закрыта!")
        if is_filled:
            messages.success(request, "Ведомость успешно заполнена!")

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = GenerateGradebookForm()  # Доп. форма
        return context


class GradebookClosedList(ObjectTableView):
    table_class = GradebookTable
    filterset_class = GradebookFilter
    queryset = Gradebook.objects.filter(status=Gradebook.STATUS_CHOICE[3][1])
    properties = {}

    def get_table_class(self):
        if self.request.GET.get("mobile") == "1":
            return GradebookMobileTable
        return GradebookTable

class TeachersGradeBookTableView(ObjectTableView):
    permission_required = None
    table_class = TeacherGradeBookTable
    filterset_class = GradeBookTeachersFilter
    queryset = Gradebook.objects.all() 

    def get_queryset(self):
        # Получаем залогиненного пользователя
        user = get_object_or_404(CustomUser, pk=self.request.user.pk)

        # Проверка аутентификации и прав преподавателя
        if not user.is_authenticated or not hasattr(user, 'is_teacher') or not user.is_teacher:
            return Gradebook.objects.none()

        # Фильтрация ведомостей
        queryset = Gradebook.objects.filter(
            status='Открыта',  # Статус "Открыта"
            teachers__id=user.id  # Текущий пользователь в списке преподавателей
        ).select_related(  # Оптимизация запросов
            'group',
            'discipline'
        ).prefetch_related(
            'teachers',
            'students'
        ).distinct()  # Убираем дубликаты

        return queryset
 
    def get_table_class(self):
        if self.request.GET.get("mobile") == "1":
            return GradebookMobileTable
        return TeacherGradeBookTable

    def get(self, request, *args, **kwargs):

        is_filled = request.GET.get('is_filled', None)

        if is_filled:
            messages.success(request, "Ведомость успешно заполнена!")

        return super().get(request, *args, **kwargs)

class GradebookDetailView(ObjectDetailView):
    """
    Класс для отображения детальной информации об учебном журнале.
    """
    model= Gradebook
    paginate_by   = 30
    template_name = 'Gradebook/detail/gradebook_student.html'

    fieldset = {
        'Основная информация': [
            'name', 'status'
        ],
    }


    def check_status(obj, user):
        return obj.status == "Не заполнена"

    buttons = [
        Button (
            id = 'change',
            name = 'Обновить',
            link_params = ['pk'],
            link_name = getpattern(Gradebook, 'change'),
            permission = getpermission(Gradebook, 'change'),
        ),
        Button (
            id = 'to_list',
            name = 'К таблице',
            link_name = getpattern(Gradebook, 'list'),
            permission = getpermission(Gradebook, 'view')
        ),
        Button (
            link_params = ['pk'],
            name = 'Открыть ведомость',
            link_name = 'check_open_gradebook',
            condition = lambda obj, user: obj.status == "Не заполнена"
        ),
        Button (
            link_params = ['pk'],
            name = 'Закрыть ведомость',
            link_name = 'download_report',
            condition = lambda obj, user: obj.status == "Заполнена"
        ),
        Button (
            link_params = ['pk'],
            name = 'Скачать документ',
            link_name = 'download_report',
            condition = lambda obj, user: obj.status == "Закрыта"
        )
    ]

    def grade_book_student_filter(self, queryset):
        if queryset.filter(gradebook_id=self.pk).count() == 0:
            return GradebookStudents.objects.filter(gradebook=self.pk)
        return queryset.filter(gradebook__id=self.pk)

    def grade_book_teacher_filter(self, queryset):
        return self.teachers.all()

    tables = [
        SubTable (
            name='Студенты',
            table=GradebookStudentsTable,
            queryset=GradebookStudents.objects.all(),
            filter_func=grade_book_student_filter,
            buttons=[
                Button (
                    name = 'Заполнить ведомость',
                    link_name = 'gradebookstudents_bulk_create',
                    link_params = ['pk']
                )
            ]
        ),
        SubTable(
            name='Учителя',
            queryset=CustomUser.objects.all(),
            table=GradebookTeachersTable,
            filter_func=grade_book_teacher_filter
        )
    ]

    def get(self, request, *args, **kwargs):
        gradebook = self.get_object()

        is_active = request.GET.get('is_active',None)

        # Проверка основных полей
        required_fields = [
            gradebook.group_id,  # Проверка ForeignKey через id
            gradebook.discipline_id,  # Проверка ForeignKey дисциплины
            gradebook.semester_number,  # Проверка IntegerField
            gradebook.status,  # Проверка CharField статуса
        ]

        # Проверка текстовых полей на непустые значения
        text_fields_valid = all([
            gradebook.status and gradebook.status.strip(),
        ])

        relations_valid = all([
            gradebook.teachers.exists(),  # Проверка наличия преподавателей
            gradebook.students.exists(),  # Проверка наличия студентов
        ])

        # Комплексная проверка всех условий
        if not all(required_fields) or not text_fields_valid or not relations_valid:
            messages.error(
                request,
                "Невозможно открыть ведомость! Заполните все обязательные поля: "
                "группа, дисциплина, название, статус, семестр, преподаватели и студенты."
            )
        else:
            if is_active:
                messages.success(request, "Ведомость успешно открыта!")

        return super().get(request, *args, **kwargs)

class GradebookUpdateView(GradeBookMixin, ObjectUpdateView):
    """
    Класс для обновления информации об учебном журнале.
    """
    form_class = GradebookForm
    queryset = Gradebook.objects.all()
    properties = {}

    buttons = [
        Button (
            id='to_object',
            name = 'К объекту',
            link_params = ['pk'],
            link_name = getpattern(Gradebook, 'detail'),
            permission = getpermission(Gradebook, 'view')
        ),
        Button (
            id = 'to_list',
            name = 'К таблице',
            link_name = getpattern(Gradebook, 'list'),
            permission = getpermission(Gradebook, 'view')
        )
    ]



class GradebookCreateView(GradeBookMixin, ObjectCreateView):
    """
    Класс для создания нового учебного журнала.
    """
    model = Gradebook
    form_class = GradebookForm
    template_name = 'Gradebook/create/grade_book.html'
    properties = ['group_id']

    buttons = [
        Button (
            id = 'to_list',
            name = 'К таблице',
            link_name = getpattern(Gradebook, 'list'),
            permission = getpermission(Gradebook, 'view')
        )
    ]


def get_download_path() -> Path:
    """
    Возвращает путь к папке загрузок в зависимости от ОС.
    Поддерживает: Windows, macOS, Linux, Android (Termux), iOS (Pythonista).
    Для других случаев возвращает текущую директорию.
    """
    # Для Android (Termux)
    if "ANDROID_ROOT" in os.environ:
        return Path("/storage/emulated/0/Download")

    # Для iOS (Pythonista)
    if sys.platform == "ios":
        return Path(os.path.expanduser("~/Documents"))

    # Для Windows
    if sys.platform == "win32":
        return Path(os.path.join(os.environ["USERPROFILE"], "Downloads"))

    # Для macOS и Linux
    download_dir = ""

    # Проверка XDG для Linux
    if sys.platform == "linux":
        download_dir = os.environ.get("XDG_DOWNLOAD_DIR", "")
        if download_dir:
            return Path(download_dir)

    # Стандартные пути для macOS/Linux/других Unix
    home = Path.home()
    possible_paths = [
        home / "Downloads",
        home / "Download",
        home / "downloads",
        home / "Загрузки"  # Для русскоязычных систем
    ]

    # Поиск существующей папки
    for path in possible_paths:
        if path.exists():
            return path

    # Fallback: текущая директория
    return Path.cwd()

@permission_required(getpermission(Gradebook, 'view'))
def download_report(request, pk):
    gradebook = get_object_or_404(Gradebook, pk=pk)

    response = redirect('gradebook_list')
    if gradebook.status == Gradebook.STATUS_CHOICE[2][1]:
        response['Location'] += '?is_closed=true'
        gradebook.status = Gradebook.STATUS_CHOICE[3][1]
        gradebook.date_of_closing = timezone.now()
        gradebook.save()

    path = get_download_path()
    print(path)
    generator = GradebookDocumentGenerator(gradebook)
    generator.generate_document(f"{path}/{gradebook.name}_{gradebook.discipline.name}_{gradebook.semester_number}_семестр_{gradebook.group.full_name}.docx")

    return response

@permission_required(getpermission(Gradebook, 'view'))
def check_and_open_gradebook(request, pk):
    """
    Проверяет заполненность всех обязательных полей ведомости,
    включая наличие связанных преподавателей и студентов.
    """
    gradebook = get_object_or_404(Gradebook, id=pk)

    # Проверка основных полей
    required_fields = [
        gradebook.group_id,  # Проверка ForeignKey через id
        gradebook.name,  # Проверка CharField
        gradebook.discipline_id,  # Проверка ForeignKey дисциплины
        gradebook.semester_number,  # Проверка IntegerField
        gradebook.status,  # Проверка CharField статуса
    ]

    # Проверка текстовых полей на непустые значения
    text_fields_valid = all([
        gradebook.name and gradebook.name.strip(),
        gradebook.status and gradebook.status.strip(),
    ])

    # Проверка связей ManyToMany и ForeignKey
    relations_valid = all([
        gradebook.teachers.exists(),  # Проверка наличия преподавателей
        gradebook.students.exists(),  # Проверка наличия студентов
    ])

    # Комплексная проверка всех условий
    if not all(required_fields) or not text_fields_valid or not relations_valid:
        return redirect('gradebook_detail', pk=pk)

    # Обновление статуса и даты
    gradebook.status = "Открыта"
    gradebook.date_of_opening = timezone.now().date()
    gradebook.save()

    response = redirect('gradebook_detail', pk=pk)
    response['Location'] += '?is_active=true'
    return response

