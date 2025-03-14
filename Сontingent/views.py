import os
import random
import re
import string

from django.utils.dateparse import parse_date
from openpyxl.reader.excel import load_workbook

from Academhub.base.mixin import ImportViewMixin
from .forms import *
from .utils import *
from .tables import *
from .filters import *
from datetime import datetime
from Academhub.models import (
    Student,
    Practice,
    TermPaper,
    Specialty,
    Gradebook,
    Discipline,
    Curriculum,
    GroupStudents,
    Qualification,
    StudentRecordBook,
    ContingentMovement,
    RecordBookTemplate,
    ProfessionalModule,
    MiddleCertification,
)
from django.conf import settings
from contextlib import nullcontext
from collections import defaultdict
from Academhub.base import SubTable
from django.http import HttpResponse
from django.utils.text import normalize_newlines
from django.urls import reverse_lazy
from Gradebook.tables import GradebookTable2
from django.shortcuts import render, get_object_or_404, redirect
from .tables import AcademTable, ExpulsionTable, ContingentMovementTable
from .filters import AcademFilter, ExpulsionFilter, ContingentMovementFilter
from .forms import AcademLeaveForm, AcademReturnForm, ExpellStudentForm, RecoverStudentForm, StudentImportForm
from Academhub.base import ObjectTableView, ObjectDetailView, ObjectUpdateView, ObjectCreateView, ObjectTemplateView
from Academhub.modules.documentGenPars import StatisticsTableGenerator, CourseTableGenerator, GroupTableGenerator, VacationTableGenerator, MovementTableGenerator

__all__ = (
    # 'view_record_book',
    'ViewRecordBookView',
    'ViewRecordBookTemplateView',
    'save_record_book_template',
    # 'create_record_book_template',
    # 'CreateRecordBookTemplateView',
    'EditRecordBookTemplateView',
    'generate_student_record_book',
    'generate_group_recordbooks',

    'DisciplineTableView',
    'DisciplineDetailView',
    'DisciplineUpdateView',
    'DisciplineCreateView',

    'SpecialtyTableView',
    'SpecialtyDetailView',
    'SpecialtyUpdateView',
    'SpecialtyCreateView',

    'StudentTableView', 
    'StudentDetailView', 
    'StudentUpdateView', 
    'StudentCreateView',

    'GroupTableView', 
    'GroupDetailView', 
    'GroupUpdateView', 
    'GroupCreateView',

    'QualificationTableView',
    'QualificationCreateView',
    'QualificationDetailView',
    'QualificationUpdateView'
)


#
## Discipline
#

class DisciplineTableView(ObjectTableView):
    """
        Класс для отображения таблицы дисциплин.
    """
    table_class = DisciplineTable
    filterset_class = DisciplineFilter
    queryset = Discipline.objects.all()

class DisciplineDetailView(ObjectDetailView):
    """
    Класс для отображения детальной информации о дисциплине.
    """
    model= Discipline
    paginate_by  = 30

    fieldset = {
        'Основная информация':
            ['name', 'code', 'specialty',]
    }
        
class DisciplineUpdateView(ObjectUpdateView):
    """
    Класс для обновления информации о дисциплине.
    """
    form_class = DisciplineForm
    queryset = Discipline.objects.all()
    
class DisciplineCreateView(ObjectCreateView):
    """
    Класс для создания новой дисциплины.
    """
    model = Discipline
    form_class = DisciplineForm

#
## Specialty
#

class SpecialtyTableView(ObjectTableView):
    """
    Класс для отображения таблицы специальностей.
    """
    table_class = SpecialtyTable
    filterset_class = SpecialtyFilter
    queryset = Specialty.objects.all()

class SpecialtyDetailView(ObjectDetailView):
    """
    Класс для отображения детальной информации о специальности.
    """
    model= Specialty
    paginate_by  = 30
    template_name = 'Contingent/detail/specialty_detail.html'

    fieldset = {
        'Основная информация':
            ['code', 'name'],
    }

    tables = [
        SubTable(
            name='Квалификации',
            filter_key='specialty',
            table=QualificationTable,
            queryset=Qualification.objects.all(),
        )
    ]

class SpecialtyUpdateView(ObjectUpdateView):
    """
    Класс для обновления информации о специальности.
    """
    form_class = SpecialtyForm
    queryset = Specialty.objects.all()

class SpecialtyCreateView(ObjectCreateView):
    """
    Класс для создания новой специальности.
    """
    model = Specialty
    form_class = SpecialtyForm

#
## Qualification
#

class QualificationTableView(ObjectTableView):
    """
    Класс для отображения таблицы квалификаций.
    """
    table_class = QualificationTable
    filterset_class = QualificationFilter
    queryset = Qualification.objects.all()

class QualificationTableView(ObjectTableView):
    """
    Класс для отображения таблицы квалификаций.
    """
    table_class = QualificationTable
    filterset_class = QualificationFilter
    queryset = Qualification.objects.all()

class QualificationDetailView(ObjectDetailView):
    """
    Класс для отображения детальной информации о квалификации.
    """
    model= Qualification
    paginate_by  = 30
    template_name = 'Contingent/detail/qualification_detail.html'

    fieldset = {
        'Основная информация':
            ['short_name', 'name', 'specialty']
    }

    tables = [
        SubTable(
            name='Студенты',
            table=GroupTable,
            filter_key='qualification',
            queryset=GroupStudents.objects.all(),
        )
    ]

class QualificationUpdateView(ObjectUpdateView):
    """
    Класс для обновления информации о квалификации.
    """
    form_class = QualificationForm
    queryset = Qualification.objects.all()

class QualificationCreateView(ObjectCreateView):
    """
    Класс для создания новой квалификации.
    """
    model = Qualification
    form_class = QualificationForm


#
## Group
#

class GroupTableView(ObjectTableView):
    """
    Класс для отображения таблицы групп.
    """
    table_class = GroupTable
    filterset_class = GroupFilter
    queryset = GroupStudents.objects.all()

class GroupDetailView(ObjectDetailView):
    """
    Класс для отображения детальной информации о группе.
    """
    paginate_by  = 30
    model= GroupStudents
    template_name = 'Contingent/detail/group_detail.html'

    fieldset = {
        'Основная информация':
            ['number', 'qualification']
    }

    tables = [
        SubTable (
            name='Студенты',
            filter_key='group',
            table=StudentTable2,
            queryset=Student.objects.filter(is_expelled=False, is_in_academ=False),
        ),
    ]

class GroupUpdateView(ObjectUpdateView):
    """
    Класс для обновления информации о группе.
    """
    form_class = GroupForm
    queryset = GroupStudents.objects.all()

class GroupCreateView(ObjectCreateView):
    """
    Класс для создания новой группы.
    """
    model = GroupStudents
    form_class = GroupForm

#
## Student
#

class StudentTableView(ImportViewMixin, ObjectTableView):
    """
    Класс для отображения таблицы студентов.
    """
    table_class = StudentTable
    filterset_class = StudentFilter
    queryset = Student.objects.filter(is_expelled=False, is_in_academ=False)
    template_name = 'Contingent/list/student_list.html'
    form_import = StudentImportForm

    def get(self, request, *args, **kwargs):
        return super().get(request, *args, **kwargs)



class StudentDetailView(ObjectDetailView):
    """
    Класс для отображения детальной информации о студенте.
    """
    model= Student
    paginate_by  = 30
    template_name = 'Contingent/detail/student_detail.html'

    fieldset = {
        'Основная информация':
            ['full_name', 'phone', 'birth_date', 'snils', 'course', 'group', 'admission_order', 'note'],
            
        'Образование':
            ['education_base', 'education_basis', 'transfer_to_2nd_year_order', 'transfer_to_3rd_year_order', 'transfer_to_4th_year_order', 'expelled_due_to_graduation', 'left_course'],
        
        'Контакты':
            ['registration_address', 'actual_address', 'representative_full_name', 'representative_email'],
    }


class StudentUpdateView(ObjectUpdateView):
    """
    Класс для обновления информации о студенте.
    """
    form_class = StudentForm
    queryset = Student.objects.all()


class StudentCreateView(ObjectCreateView):
    """
    Класс для создания нового студента.
    """
    model = Student
    form_class = StudentForm

#
##
#

class AcademUpdateView(ObjectUpdateView):
    """
    Класс для отправки студента в академ.
    """
    form_class = AcademLeaveForm
    queryset = Student.objects.all()

class AcademListView(ObjectTableView):
    """
    Класс для просмотра студентов, находящихся в академе
    """
    table_class = AcademTable
    filterset_class = AcademFilter
    queryset = Student.objects.filter(is_in_academ=True)
    template_name = 'Contingent/academ_list.html'

class AcademReturn(ObjectUpdateView):
    form_class = AcademReturnForm
    queryset = Student.objects.all()


class ExpulsionListView(ObjectTableView):
    table_class = ExpulsionTable
    filterset_class = ExpulsionFilter
    queryset = Student.objects.filter(is_expelled=True)

class ExpelStudent(ObjectUpdateView):
    form_class = ExpellStudentForm
    queryset = Student.objects.all()

class RecoverStudent(ObjectUpdateView):
    form_class = RecoverStudentForm
    queryset = Student.objects.filter(is_expelled=True)

class StatisticksView(ObjectTemplateView):
    template_name = 'Contingent/statisticks.html'

    def get_context_data(self, **kwargs):
        student_list = student_format_to_list()
        specialty_data = defaultdict(lambda: defaultdict(int))
        all_specialties = list(Specialty.objects.all().values_list())
        print(all_specialties)

        # Инициализируем все направления заранее
        for qualification in Qualification.objects.all():
            specialty_data[(
            qualification.specialty.code, qualification.specialty.name, qualification.name)]  # Создаем пустые записи
            all_specialties = [spec for spec in all_specialties if spec[0] != qualification.specialty.id]

        for specialty in all_specialties:
            specialty_data[(specialty[1], specialty[2], "")]

        for student in student_list:
            spec_code = student["specialty_code"]
            spec_name = student["specialty_name"]
            qualification = student["qualification"]
            course = student["course"]
            base = student["base"]  # Основное общее или Среднее общее
            budget = student["budget"]  # Бюджет или Внебюджет
            academic = student["academic_leave"]  # True - в академе, False - нет
            expelled = student["is_expelled"]

            if expelled:
                continue

            if base == "Основное общее":
                base = 9
            else:
                base = 11

            if budget == "Бюджет":
                budget = True
            else:
                budget = False

            if academic:
                key = f"students_academic_{course}_{'budget' if budget else 'paid'}"
            else:
                key = f"students_{base}_{course}_{'budget' if budget else 'paid'}"

            specialty_data[(spec_code, spec_name, qualification)][key] += 1

        number = 1
        table_data = []
        total_9_1_ = 0
        total_9_2_ = 0
        total_9_3_ = 0
        total_9_4_ = 0

        total_11_1_ = 0
        total_11_2_ = 0
        total_11_3_ = 0

        total_academ_1 = 0
        total_academ_2 = 0
        total_academ_3 = 0
        total_academ_4 = 0

        total_contingent = 0

        for (code, name, qualification), counts in specialty_data.items():
            total_budget = sum(counts.get(key, 0) for key in counts if "budget" in key)
            total_paid = sum(counts.get(key, 0) for key in counts if "paid" in key)

            total_9_1_ += sum(counts.get(key, 0) for key in counts if "9_1" in key)
            total_9_2_ += sum(counts.get(key, 0) for key in counts if "9_2" in key)
            total_9_3_ += sum(counts.get(key, 0) for key in counts if "9_3" in key)
            total_9_4_ += sum(counts.get(key, 0) for key in counts if "9_4" in key)

            total_11_1_ += sum(counts.get(key, 0) for key in counts if "11_1" in key)
            total_11_2_ += sum(counts.get(key, 0) for key in counts if "11_2" in key)
            total_11_3_ += sum(counts.get(key, 0) for key in counts if "11_3" in key)

            total_academ_1 += sum(counts.get(key, 0) for key in counts if "academic_1" in key)
            total_academ_2 += sum(counts.get(key, 0) for key in counts if "academic_2" in key)
            total_academ_3 += sum(counts.get(key, 0) for key in counts if "academic_3" in key)
            total_academ_4 += sum(counts.get(key, 0) for key in counts if "academic_4" in key)

            total_contingent += total_budget + total_paid

            row = [
                number,
                code,  # Код специальности
                name,  # Название специальности
                qualification,  # Квалификация
                counts.get("students_9_1_budget", 0),
                counts.get("students_9_1_paid", 0),
                counts.get("students_9_2_budget", 0),
                counts.get("students_9_2_paid", 0),
                counts.get("students_9_3_budget", 0),
                counts.get("students_9_3_paid", 0),
                counts.get("students_9_4_budget", 0),
                counts.get("students_9_4_paid", 0),
                counts.get("students_11_1_budget", 0),
                counts.get("students_11_1_paid", 0),
                counts.get("students_11_2_budget", 0),
                counts.get("students_11_2_paid", 0),
                counts.get("students_11_3_budget", 0),
                counts.get("students_11_3_paid", 0),
                counts.get("students_academic_1_budget", 0),
                counts.get("students_academic_1_paid", 0),
                counts.get("students_academic_2_budget", 0),
                counts.get("students_academic_2_paid", 0),
                counts.get("students_academic_3_budget", 0),
                counts.get("students_academic_3_paid", 0),
                counts.get("students_academic_4_budget", 0),
                counts.get("students_academic_4_paid", 0),
                total_budget,  # Итого бюджет
                total_paid,  # Итого внебюджет
            ]
            table_data.append(row)
            number += 1

        last_row = [total_9_1_, total_9_2_, total_9_3_, total_9_4_, total_11_1_, total_11_2_, total_11_3_,
                    total_academ_1,
                    total_academ_2, total_academ_3, total_academ_4, total_contingent]

        total_1_09_02_07_budget = sum(
            counts.get(key, 0)
            for (code, _, _), counts in specialty_data.items()
            if code == "09.02.07"
            for key in counts
            if "budget" in key and key.startswith("students_9_1")
        )

        total_1_09_02_07_paid = sum(
            counts.get(key, 0)
            for (code, _, _), counts in specialty_data.items()
            if code == "09.02.07"
            for key in counts
            if "paid" in key and key.startswith("students_9_1")
        )

        context = super().get_context_data(**kwargs)
        context = context|{'table_data': table_data, 'last_row': last_row,
                   'total_09_02_07_budget': total_1_09_02_07_budget,
                   'total_09_02_07_paid': total_1_09_02_07_paid}

        return context


def save_record_book_template(request, qualification_id, admission_year):
    """
    Функция для сохранения ШАБЛОНА зачетной книжки
    """
    qualification = get_object_or_404(Qualification, id=qualification_id)
    curriculum = get_object_or_404(Curriculum, qualification=qualification, admission_year=admission_year)
    template = RecordBookTemplate.objects.get(qualification=qualification, admission_year=admission_year)

    if request.method == 'POST':
        # Обновление основной информации
        template.student_name = request.POST.get('student_name', '')
        template.record_book_number = request.POST.get('record_book_number', '')
        template.admission_order = request.POST.get('admission_order', '')
        template.issue_date = request.POST.get('issue_date')
        template.save()

        # Обработка промежуточных аттестаций
        middle_semesters = request.POST.getlist('middle_semester[]')
        middle_disciplines = request.POST.getlist('middle_discipline[]')
        middle_hours = request.POST.getlist('middle_hours[]')
        middle_is_exams = request.POST.getlist('middle_is_exam[]')
        template.middle_certifications.clear()
        for sem, disc, hrs, is_exam in zip(middle_semesters, middle_disciplines, middle_hours, middle_is_exams):
            cert = MiddleCertification.objects.create(
                semester=sem,
                discipline_id=disc,
                hours=hrs,
                is_exam=is_exam == 'True'
            )
            template.middle_certifications.add(cert)

        # Обработка модулей
        module_names = request.POST.getlist('module_name[]')
        module_hours = request.POST.getlist('module_hours[]')
        template.professional_modules.clear()
        for name, hrs in zip(module_names, module_hours):
            module = ProfessionalModule.objects.create(module_name=name, hours=hrs)
            template.professional_modules.add(module)

        # Обработка практик
        practice_names = request.POST.getlist('practice_name[]')
        practice_hours = request.POST.getlist('practice_hours[]')
        practice_semesters = request.POST.getlist('practice_semester[]')
        template.practices.clear()
        for name, hrs, sem in zip(practice_names, practice_hours, practice_semesters):
            practice = Practice.objects.create(practice_name=name, hours=hrs, semester=sem, practice_type='УП')
            template.practices.add(practice)

        # Обработка курсовых
        term_disciplines = request.POST.getlist('term_discipline[]')
        term_topics = request.POST.getlist('term_topic[]')
        term_grades = request.POST.getlist('term_grade[]')
        template.term_papers.clear()
        for disc, topic, grade in zip(term_disciplines, term_topics, term_grades):
            paper = TermPaper.objects.create(discipline_id=disc, topic=topic, grade=grade)
            template.term_papers.add(paper)

        # return redirect('qualification_detail', pk=qualification_id)
        return redirect('view_record_book_template', qualification_id=qualification_id, admission_year=admission_year)

    return redirect('create_record_book_template', qualification_id=qualification_id, admission_year=admission_year)


class ViewRecordBookTemplateView(ObjectTemplateView):
    """
    Класс для отображения ШАБЛОНА зачетной книжки.
    """
    template_name = 'Contingent/record_book_view.html'
    model = RecordBookTemplate

    def get_context_data(self, **kwargs):
        # Получаем базовый контекст от родительских классов
        context = super().get_context_data(**kwargs)

        qualification = get_object_or_404(Qualification, id=self.kwargs['qualification_id'])
        template = get_object_or_404(RecordBookTemplate, qualification=qualification,
                                     admission_year=self.kwargs['admission_year'])

        context['qualification'] = qualification
        context['template'] = template
        context['admission_year'] = self.kwargs['admission_year']
        context['url_list'] = 'qualification_list'

        return context

class ViewRecordBookView(ObjectTemplateView):
    """
    Класс для отображения информации о зачетной книжке конкретного студента.
    """
    template_name = 'Contingent/record_book_view.html'
    model = StudentRecordBook

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qualification = get_object_or_404(Qualification, id=self.kwargs['qualification_id'])
        template = get_object_or_404(StudentRecordBook, student=self.kwargs['student_id'])
        context['qualification'] = qualification

        context['qualification'] = qualification
        context['template'] = template
        context['admission_year'] = self.kwargs['admission_year']
        context['url_list'] = 'student_list'
        context['is_student_gradebbok'] = True
        context['student_id'] = self.kwargs['student_id']
        return context

class EditRecordBookTemplateView(ObjectUpdateView):
    """
    редактирование Шаблоа зачетки студентов
    """
    model = RecordBookTemplate
    template_name = 'Contingent/record_book_edit.html'
    fields = ['student_name', 'record_book_number', 'admission_order', 'issue_date']

    def get_object(self, queryset=None):
        qualification = get_object_or_404(Qualification, id=self.kwargs['qualification_id'])
        return get_object_or_404(RecordBookTemplate, qualification=qualification, admission_year=self.kwargs['admission_year'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qualification = get_object_or_404(Qualification, id=self.kwargs['qualification_id'])
        context['qualification'] = qualification
        context['admission_year'] = self.kwargs['admission_year']
        context['curriculum'] = self.object.curriculum
        context['curriculum_disciplines'] = [{"id": d.id, "name": d.name} for d in self.object.curriculum.disciplines.all()]
        context['url_list'] = 'qualification_list'
        context['mobel_verbosename'] = self.get_verbose_name()  # Совместимость с ObjectUpdateView
        return context

    def form_valid(self, form):
        self.object = form.save()

        # Обработка дополнительных данных формы
        middle_semesters = self.request.POST.getlist('middle_semester[]')
        middle_disciplines = self.request.POST.getlist('middle_discipline[]')
        middle_hours = self.request.POST.getlist('middle_hours[]')
        middle_is_exams = self.request.POST.getlist('middle_is_exam[]')
        self.object.middle_certifications.clear()
        for sem, disc, hrs, is_exam in zip(middle_semesters, middle_disciplines, middle_hours, middle_is_exams):
            cert = MiddleCertification.objects.create(
                semester=sem,
                discipline_id=disc,
                hours=hrs,
                is_exam=is_exam == 'True'
            )
            self.object.middle_certifications.add(cert)

        module_names = self.request.POST.getlist('module_name[]')
        module_hours = self.request.POST.getlist('module_hours[]')
        self.object.professional_modules.clear()
        for name, hrs in zip(module_names, module_hours):
            module = ProfessionalModule.objects.create(module_name=name, hours=hrs)
            self.object.professional_modules.add(module)

        practice_names = self.request.POST.getlist('practice_name[]')
        practice_hours = self.request.POST.getlist('practice_hours[]')
        practice_semesters = self.request.POST.getlist('practice_semester[]')
        self.object.practices.clear()
        for name, hrs, sem in zip(practice_names, practice_hours, practice_semesters):
            practice = Practice.objects.create(practice_name=name, hours=hrs, semester=sem, practice_type='УП')
            self.object.practices.add(practice)

        term_disciplines = self.request.POST.getlist('term_discipline[]')
        self.object.term_papers.clear()
        for disc in term_disciplines:
            paper = TermPaper.objects.create(discipline_id=disc)
            self.object.term_papers.add(paper)

        return redirect('view_record_book', qualification_id=self.kwargs['qualification_id'], admission_year=self.kwargs['admission_year'])

    def get_success_url(self):
        return reverse_lazy('view_record_book_template', kwargs={'qualification_id': self.kwargs['qualification_id'], 'admission_year': self.kwargs['admission_year']})


def generate_student_record_book(request, pk):
    """
    Функция для генерации зачетной книжки для конкретного студента
    """
    try:
        student = get_object_or_404(Student, pk=pk)
    except StudentRecordBook.DoesNotExist as e:
        print("Студент не найден")
    qualification = student.group.qualification
    admission_year = student.group.year_create

    print(qualification, admission_year)
    try:
        template = get_object_or_404(RecordBookTemplate, qualification=qualification, admission_year=admission_year)
    except RecordBookTemplate.DoesNotExist as e:
        print("Шаблон не найден")

    # Генерация уникального номера зачетки
    record_book_number = generate_unique_record_book_number(admission_year, student)
    print(2)
    # Создаём новый объект StudentRecordBook для студента
    new_record_book = StudentRecordBook.objects.create(
        student=student,
        qualification=qualification,
        admission_year=admission_year,
        student_name=f"{student.full_name}",
        record_book_number=record_book_number,
        admission_order=template.admission_order,
        issue_date=template.issue_date,
        curriculum=template.curriculum
    )
    print(new_record_book)
    # Копируем ManyToMany-поля из шаблона
    new_record_book.middle_certifications.set(template.middle_certifications.all())
    new_record_book.professional_modules.set(template.professional_modules.all())
    new_record_book.practices.set(template.practices.all())
    new_record_book.term_papers.set(template.term_papers.all())

    # Связываем зачётку со студентом
    student.record_book = new_record_book
    student.save()

    return redirect('view_record_book', qualification_id=qualification.id, admission_year=admission_year, student_id=pk)


def create_auto_record_book_template(request, qualification_id, admission_year):
    """
    Функция для генерации Шаблона зачетной книжки  шаблон генерируется на квалификацию и год поступления,
    Например для всех групп П 2021 года поступления
    """
    qualification = get_object_or_404(Qualification, id=qualification_id)
    curriculum = get_object_or_404(Curriculum, qualification=qualification, admission_year=admission_year)

    # Проверяем, существует ли шаблон, чтобы не дублировать
    if RecordBookTemplate.objects.filter(qualification=qualification, admission_year=admission_year).exists():
        return redirect('view_record_book_template', qualification_id=qualification_id, admission_year=admission_year)

    # Создаём шаблон
    template = RecordBookTemplate.objects.create(
        qualification=qualification,
        admission_year=admission_year,
        student_name="",  # Пустое, так как это шаблон
        record_book_number="",  # Пустое для шаблона
        admission_order="Не указан",  # Можно заполнить позже
        issue_date=datetime.now().date(),
        curriculum=curriculum
    )

    # Автоматически заполняем компоненты зачётки из CurriculumItem
    for item in curriculum.items.all():
        if item.item_type == 'discipline' and item.discipline:
            # Создаём MiddleCertification для дисциплин
            middle_cert = MiddleCertification.objects.create(
                semester=item.semester,
                discipline=item.discipline,
                hours=item.hours,
                is_exam=(item.attestation_form == 'exam')
            )
            template.middle_certifications.add(middle_cert)
        elif item.item_type == 'practice' and item.practice:
            # Связываем существующую практику
            template.practices.add(item.practice)
        elif item.item_type == 'professional_module' and item.professional_module:
            # Связываем существующий модуль
            template.professional_modules.add(item.professional_module)
        elif item.item_type == 'term_paper' and item.term_paper:
            # Связываем существующую курсовую
            template.term_papers.add(item.term_paper)

    return redirect('view_record_book_template', qualification_id=qualification_id, admission_year=admission_year)


def generate_group_recordbooks(request, group_id):
    """
    Функция для генерации зачетных книжек на всю группу
    """
    group = get_object_or_404(GroupStudents, id=group_id)
    qualification = group.qualification
    admission_year = group.year_create

    # проверка на то что шаблон есть
    if not RecordBookTemplate.objects.filter(qualification=qualification, admission_year=admission_year).exists():
        create_auto_record_book_template(request, qualification.id, admission_year)

    template = get_object_or_404(RecordBookTemplate, qualification=qualification, admission_year=admission_year)
    students = group.students.all()

    for student in students:
        if student.record_book:  # Пропускаем студентов с уже существующей зачёткой
            continue

        record_book_number = generate_unique_record_book_number(admission_year, student)
        new_record_book = StudentRecordBook.objects.create(
            student=student,
            qualification=qualification,
            admission_year=admission_year,
            student_name=f"{student.full_name}",
            record_book_number=record_book_number,
            admission_order=template.admission_order,
            issue_date=template.issue_date,
            curriculum=template.curriculum
        )

        # Копируем данные из шаблона
        new_record_book.middle_certifications.set(template.middle_certifications.all())
        new_record_book.professional_modules.set(template.professional_modules.all())
        new_record_book.practices.set(template.practices.all())
        new_record_book.term_papers.set(template.term_papers.all())

        student.record_book = new_record_book
        student.save()

    return redirect('groupstudents_detail', pk=group_id)


class ContingentMovementTableView(ObjectTableView):
    """
    Класс для отображения таблицы движений контингента.
    """
    table_class = ContingentMovementTable
    filterset_class = ContingentMovementFilter
    queryset = ContingentMovement.objects.all()
    # template_name = 'base_view.html'
    template_name = 'Contingent/contingent_movement_list.html'
    paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['table_name'] = "Движения контингента"
        return context


#########
#######Обработка генерации доков
#########


def generate_group_table(request):
    """
    Функция для генерации файла .xslx для всех групп
    """
    groups = GroupStudents.objects.all()
    generator = GroupTableGenerator(groups)
    file_path = os.path.join(settings.MEDIA_ROOT, 'group_table.xlsx')

    # Создаём директорию, если она не существует
    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

    generator.generate_document(file_path)

    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="group_table.xlsx"'

    os.remove(file_path)  # Удаляем временный файл после отправки
    return response


def generate_course_table(request, course):
    """
    Функция для генерации файла .xslx для статистики по определенному курсу
    """
    students = Student.objects.filter(group__current_course=course)
    try:
        generator = CourseTableGenerator(students)
        file_path = os.path.join(settings.MEDIA_ROOT, f'course_{course}_table.xlsx')

        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

        generator.generate_document(file_path)

        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="course_{course}_table.xlsx"'
        os.remove(file_path)
        return response
    except Exception as e:
        return HttpResponse(f"Ошибка: {e.message}", status=400)


def generate_statistics_table(request):
    """
    Функция для генерации файла .xslx для статистики
    """
    specialties = Specialty.objects.all()
    qualifications = Qualification.objects.all()
    students = Student.objects.all()
    generator = StatisticsTableGenerator(specialties, qualifications, students)
    file_path = os.path.join(settings.MEDIA_ROOT, 'statistics_table.xlsx')

    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

    generator.generate_document(file_path)

    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="statistics_table.xlsx"'
    os.remove(file_path)
    return response


def generate_vacation_table(request):
    """
    Функция для генерации файла .xslx для студентов находящихся в академическом отпуске
    """
    students = Student.objects.filter(is_in_academ=True)
    generator = VacationTableGenerator(students)
    file_path = os.path.join(settings.MEDIA_ROOT, 'vacation_table.xlsx')

    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

    generator.generate_document(file_path)

    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="vacation_table.xlsx"'
    os.remove(file_path)
    return response


def generate_movement_table(request):
    movements = ContingentMovement.objects.all()  # Получаем все записи о движениях
    generator = MovementTableGenerator(movements)
    file_path = os.path.join(settings.MEDIA_ROOT, 'movement_table.xlsx')

    os.makedirs(settings.MEDIA_ROOT, exist_ok=True)

    generator.generate_document(file_path)

    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="movement_table.xlsx"'
    os.remove(file_path)
    return response


def import_students(request):
    """
    Функция для импорта контингентов из файла форматат .xlsx, выгруженного из 1с приемной комиссией
    """
    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active

            # Сопоставление заголовков с индексами столбцов
            headers = {cell.value: idx for idx, cell in enumerate(ws[1]) if cell.value}

            # Пропускаем строку заголовков и начинаем с данных
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    # Извлекаем данные из строки
                    last_name = row[headers.get("Фамилия")]
                    first_name = row[headers.get("Имя")]
                    middle_name = row[headers.get("Отчество")]
                    full_name = f"{last_name} {first_name} {middle_name}".strip() if last_name and first_name else None
                    group_number = row[headers.get("Академическая группа")]
                    course_str = row[headers.get("Курс")]
                    course = int(course_str[0]) if course_str and course_str[0].isdigit() else None
                    education_basis = row[headers.get("Основы обучения")]
                    birth_date = (
                        row[headers.get("Дата рождения")].date() if isinstance(row[headers.get("Дата рождения")],datetime)
                        else datetime.strptime(row[headers.get("Дата рождения")], "%d.%m.%Y").date()) if row[
                        headers.get("Дата рождения")] else None

                    phone = row[headers.get("Телефон мобильный")]
                    admission_order = row[headers.get("Приказ о зачислении")]
                    expell_order = row[headers.get("Приказ об отчислении")] if row[
                        headers.get("Приказ об отчислении")] else None
                    date_of_expelling = (
                        row[headers.get("Приказ об отчислении дата ОТ")].date() if isinstance(row[headers.get("Приказ об отчислении дата ОТ")], datetime)
                        else datetime.strptime(row[headers.get("Приказ об отчислении дата ОТ")], "%d.%m.%Y").date()) if row[
                        headers.get("Приказ об отчислении дата ОТ")] else None
                    academ_leave_date = parse_date(str(row[headers.get("Дата начала последнего академ отпуска")])) if \
                    row[headers.get("Дата начала последнего академ отпуска")] else None
                    academ_return_date = parse_date(
                        str(row[headers.get("Дата окончания последнего академ отпуска")])) if row[
                        headers.get("Дата окончания последнего академ отпуска")] else None
                    registration_addres = row[headers.get("Адрес по прописке")] if row[headers.get("Адрес по прописке")] else None
                    actual_addres = row[headers.get("Адрес проживания")] if row[headers.get("Адрес проживания")] else None
                    snils = row[headers.get("СПС")] if row[headers.get("СПС")] else None
                    ancete_number = extract_application_number(row[headers.get("Анкета абитуриента")])
                    expelled_due_to_graduation = False
                    is_expelled = False
                    reason_of_expelling = None
                    note = None
                    if expell_order:
                        is_expelled = True
                        if 'окончании' in expell_order:
                            expelled_due_to_graduation = True
                            reason_of_expelling = "Окончание обучения"
                            note = 'Отчислен в связи с окончанием обучения'

                    # Проверяем и создаём группу, если её нет
                    group = GroupStudents.objects.get(full_name=group_number)

                    # Создаём или обновляем студента
                    student, created = Student.objects.update_or_create(
                        full_name=full_name,
                        defaults={
                            'birth_date': birth_date,
                            'group': group,
                            'education_basis': education_basis,
                            'phone': phone,
                            'admission_order': admission_order,
                            'expell_order': expell_order,
                            'date_of_expelling': date_of_expelling,
                            'is_in_academ': bool(academ_leave_date and not academ_return_date),
                            'academ_leave_date': academ_leave_date,
                            'academ_return_date': academ_return_date,
                            'registration_address': registration_addres,
                            'actual_address': actual_addres,
                            'snils': snils,
                            'expelled_due_to_graduation': expelled_due_to_graduation,
                            'reason_of_expelling': reason_of_expelling,
                            'is_expelled': is_expelled,
                            'note': note,
                            'ancete_number': ancete_number,
                        }
                    )
                except Exception as e:
                    return HttpResponse(f"Ошибка при импорте строки: {str(e)}", status=400)

            return HttpResponse("Импорт студентов успешно завершён!")
    else:
        form = StudentImportForm()

    return render(request, 'Contingent/import_students.html', {'form': form})
