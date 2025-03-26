import os
from cProfile import label

from .utils import *
from django import forms
from datetime import datetime
from Academhub.forms import *
from django.utils.dateparse import parse_date
from openpyxl.reader.excel import load_workbook
from django.core.exceptions import ValidationError
from Academhub.models import Student, Discipline, GroupStudents, Specialty, Qualification

__all__ = [
    'GroupForm',
    'StudentForm',
    'SpecialtyForm',
    'DisciplineForm', 
    'StudentImportForm',
    'QualificationForm',
    'PromoteGroupStudentsForm',
    'PromoteStudentsForm'
]

class DisciplineForm(forms.ModelForm):
    specialty = forms.ModelChoiceField(
        queryset=Specialty.objects.all(),
        label='Специальность',
    )

    class Meta:
        model = Discipline
        fields = [
            'name',
            'code',
            'specialty',
        ]


class PromoteGroupStudentsForm(forms.ModelForm):
    """
    Форма для перевода группы и студентов на следующий курс
    """
    transfer_order = forms.CharField(
        label='Номер приказа о переводе',
        max_length=255,
        required=True,
    )

    class Meta:
        model = GroupStudents
        fields = []  # Пустой список, так как редактируем только transfer_order

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['transfer_order'].widget.attrs.update({
            'class': 'form-control',
            'autocomplete': 'off'
        })


class StudentForm(forms.ModelForm):
    birth_date = forms.DateField(
        label='Дата рождения',
        widget=forms.DateInput(attrs={'type': 'date'})
    )

    full_name = forms.CharField(
        label='ФИО',
        widget=forms.TextInput(
            attrs={
                'placeholder': 'Иванов Иван Иванович'
            }
        )
    )

    representative_full_name = forms.CharField(
        label='ФИО представителя',
        widget=forms.TextInput(
            attrs={
                'placeholder': 'Иванов Иван Иванович'
            }
        )
    )

    phone = forms.CharField(
        label='Телефон',
        widget=Phone(attrs={'class': 'phone-input'})
    )

    snils = forms.CharField(
        label='Снилс',
        max_length=14,
        widget=Snils(attrs={
            'class': 'delete-arrow-input-number'
        })
    )

    representative_email = forms.EmailField(
        label='Почта представителя',
        widget=forms.EmailInput(
            attrs={
                'placeholder': 'email@email.email'
            }
        )
    )

    class Meta:
        model = CurrentStudent
        fields = [
            'full_name',
            'phone',
            'birth_date',
            'snils',
            'admission_order',
            'education_base',
            'education_basis',
            'group',
            'registration_address',
            'actual_address',
            'representative_full_name',
            'representative_email',
            'note',
        ]

    def clean(self):
        cleaned_data = super().clean()
        group = cleaned_data.get('group')
        education_base = cleaned_data.get('education_base')

        # :TODO нада что-то придумать для импорта, потому что в доке может быть как 'Основное общее' так и 'основное общее'
        # if group and education_base and education_base != group.education_base:
        #     raise forms.ValidationError(
        #         f"База образования студента ({education_base}) должна совпадать с "
        #         f"базой образования группы ({group.education_base})"
        #     )
        return cleaned_data

class GroupForm(forms.ModelForm):
    class Meta:
        model = GroupStudents
        fields = [
            'qualification',
            'year_create',
            'education_base',
            'current_course',
        ]

class QualificationForm(forms.ModelForm):
    class Meta:
        model = Qualification
        fields = '__all__'

class SpecialtyForm(forms.ModelForm ):
    class Meta:
        model = Specialty
        fields = '__all__'

class AcademLeaveForm(forms.ModelForm):

    class Meta:
        model = CurrentStudent
        fields = [
            'academ_leave_date',
            'academ_return_date',
            'reason_of_academ',
            'academ_order'
        ]
        widgets = {
            'academ_leave_date': forms.DateInput(
                attrs={'type': 'date'}, format='%Y-%m-%d'
            ),
            'academ_return_date': forms.DateInput(
                attrs={'type': 'date'}, format='%Y-%m-%d'
            ),
        }

    def save(self, commit=True):
        student = super().save(commit=False)  # Получаем объект студента, но пока не сохраняем
        student.is_in_academ = True  # Переводим в академ
        student.expelled_due_to_graduation = False
        student.left_course = student.group.current_course  # Берем текущий курс группы

        if commit:
            student.save()  # Сохраняем изменения
        return student

class AcademReturnForm(forms.ModelForm):

    class Meta:
        model = Student
        fields = []

    def save(self, commit=True):
        student = super().save(commit=False)  # Получаем объект студента, но пока не сохраняем
        student.is_in_academ = False
        student.academ_leave_date = None
        student.academ_return_date = None
        student.reason_of_academ = None
        student.left_course = None
        student.expelled_due_to_graduation = False
        student.academ_order = ""
        student.save()

        if commit:
            student.save()  # Сохраняем изменения
        return student

class ExpellStudentForm(forms.ModelForm):
    class Meta:
        model = Student
        fields = [
            'date_of_expelling',
            'reason_of_expelling',
            'expell_order',
        ]

        widgets = {
            'date_of_expelling': forms.DateInput(
                attrs={'type': 'date'}, format='%Y-%m-%d'
            )
        }

    def save(self, commit=True):
        student = super().save(commit=False)  # Получаем объект студента, но пока не сохраняем
        student.is_expelled = True
        student.left_course = student.group.current_course  # Берем текущий курс группы
        student.academ_leave_date = None
        student.academ_return_date = None
        student.reason_of_academ = None
        student.is_in_academ = False
        student.expelled_due_to_graduation = False
        if commit:
            student.save()  # Сохраняем изменения
        return student

class RecoverStudentForm(forms.ModelForm):
    class Meta:
        model = Student
        fields = [
            'group',
            'reinstaitment_order'
        ]

    def save(self, commit=True):
        student = super().save(commit=False)
        student.is_expelled = False
        student.date_of_expelling = None
        student.reason_of_expelling = None
        student.left_course = None
        student.expelled_due_to_graduation = False
        if commit:
            student.save()
        return student


class StudentImportForm(forms.Form):
    excel_file = forms.FileField(
        label="Выберите файл Excel (.xlsx)"
    )


    def clean(self):
        cleaned_data = super().clean()
        file = cleaned_data['excel_file']

        if file:
            file_extension = os.path.splitext(file.discipline_name)[1]

            if file_extension.lower() != '.xlsx':
                raise ValidationError("Файл должен быть в формате .xlsx")
        
            self.get_data_from_file(file)

        return file

    def get_data_from_file(self, file):
            wb = load_workbook(file)
            ws = wb.active

            headers = {
                cell.value: idx for idx, cell in enumerate(ws[1]) if cell.value
            }

            self.data_list = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    return  # Пропускаем пустые строки

                data = {}  # Создаем словарь для хранения данных текущей строки

                last_name = row[headers.get("Фамилия")]
                first_name = row[headers.get("Имя")]
                middle_name = row[headers.get("Отчество")]
                data['full_name'] = f"{last_name} {first_name} {middle_name}".strip() if last_name and first_name else None
                data['group_number'] = row[headers.get("Академическая группа")]
                course_str = row[headers.get("Курс")]
                data['course'] = int(course_str[0]) if course_str and course_str[0].isdigit() else None
                data['education_basis'] = row[headers.get("Основы обучения")]
                data['birth_date'] = (
                    row[headers.get("Дата рождения")].date() if isinstance(row[headers.get("Дата рождения")], datetime)
                    else datetime.strptime(row[headers.get("Дата рождения")], "%d.%m.%Y").date()) if row[
                    headers.get("Дата рождения")] else None

                data['phone'] = row[headers.get("Телефон мобильный")]
                data['admission_order'] = row[headers.get("Приказ о зачислении")]
                data['expell_order'] = row[headers.get("Приказ об отчислении")] if row[
                    headers.get("Приказ об отчислении")] else None
                data['date_of_expelling'] = (
                    row[headers.get("Приказ об отчислении дата ОТ")].date() if isinstance(row[headers.get("Приказ об отчислении дата ОТ")], datetime)
                    else datetime.strptime(row[headers.get("Приказ об отчислении дата ОТ")], "%d.%m.%Y").date()) if row[
                    headers.get("Приказ об отчислении дата ОТ")] else None
                data['academ_leave_date'] = parse_date(str(row[headers.get("Дата начала последнего академ отпуска")])) if \
                row[headers.get("Дата начала последнего академ отпуска")] else None
                data['academ_return_date'] = parse_date(
                    str(row[headers.get("Дата окончания последнего академ отпуска")])) if row[
                    headers.get("Дата окончания последнего академ отпуска")] else None
                data['registration_addres'] = row[headers.get("Адрес по прописке")] if row[headers.get("Адрес по прописке")] else None
                data['actual_addres'] = row[headers.get("Адрес проживания")] if row[headers.get("Адрес проживания")] else None
                data['snils'] = row[headers.get("СПС")] if row[headers.get("СПС")] else None
                data['ancete_number'] = extract_application_number(row[headers.get("Анкета абитуриента")])
                data['expelled_due_to_graduation'] = False
                data['is_expelled'] = False
                data['reason_of_expelling'] = None
                data['note'] = None

                if data['expell_order']:
                    data['is_expelled'] = True
                    if 'окончании' in data['expell_order']:
                        data['expelled_due_to_graduation'] = True
                        data['reason_of_expelling'] = "Окончание обучения"
                        data['note'] = 'Отчислен в связи с окончанием обучения'

                # Проверяем группу
                group_filter = GroupStudents.objects.filter(
                    full_name=data['group_number']
                )

                if group_filter.exists():
                    data['group'] = group_filter.first()
                else:
                    raise ValidationError(
                        f'Группы {data["group_number"]} не существует'
                    )

                self.data_list.append(data)  # Добавляем словарь в массив

    def save(self):
        for data in self.data_list:  # Перебираем массив данных
            # Создаём или обновляем студента
            student, created = Student.objects.update_or_create(
                full_name=data['full_name'],
                defaults={
                    'birth_date': data['birth_date'],
                    'group': data['group'],
                    'education_basis': data['education_basis'],
                    'phone': data['phone'],
                    'admission_order': data['admission_order'],
                    'expell_order': data['expell_order'],
                    'date_of_expelling': data['date_of_expelling'],
                    'is_in_academ': bool(data['academ_leave_date'] and not data['academ_return_date']),
                    'academ_leave_date': data['academ_leave_date'],
                    'academ_return_date': data['academ_return_date'],
                    'registration_address': data['registration_addres'],
                    'actual_address': data['actual_addres'],
                    'snils': data['snils'],
                    'expelled_due_to_graduation': data['expelled_due_to_graduation'],
                    'reason_of_expelling': data['reason_of_expelling'],
                    'is_expelled': data['is_expelled'],
                    'note': data['note'],
                    'ancete_number': data['ancete_number'],
                }
        )


class ContingentStudentImportForm(StudentImportForm):
    def get_data_from_file(self, file):
        wb = load_workbook(file)
        # print(f"Листы в файле: {wb.sheetnames}")  # Проверяем список листов

        # sheet_name = "1 курс (9 кл)"  # Строго задаем имя листа
        required_sheets = [
            "1 курс (9 кл)",
            '1 курс (11 кл)',
            '2 курс (9 кл)',
            '3 курс (9кл)',
            '2 курс (11 кл) ',
            '4 курс (9 кл.)  ',
            '3 курс (11 кл)',
            # 'Статистика',
            # 'Движение ',
            # 'Академический отпуск'
        ]
        if required_sheets not in wb.sheetnames:
            # print(f"Лист {required_sheets} не найдены!")
            return

        for sheet_name in wb.sheetnames:
            if sheet_name in required_sheets:
                # if sheet_name == required_sheets[0] or sheet_name == required_sheets[1] or sheet_name == required_sheets[2] \
                #         or sheet_name == required_sheets[3] or sheet_name == required_sheets[4] or sheet_name == required_sheets[5] \
                #         or sheet_name == required_sheets[6]:
                ws = wb[sheet_name]  # Получаем лист

                # Заголовки
                headers = {cell.value: idx for idx, cell in enumerate(ws[1]) if cell.value}
                # print(f"Заголовки в файле: {headers}")  # Проверяем заголовки

                self.data_list = []

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(cell is None for cell in row):
                        continue  # Просто пропускаем пустые строки

                    data = {
                        'full_name': row[headers.get("ФИО")],
                        'birth_date': None,
                        'group_number': row[headers.get("Группа")],
                        'education_base': row[headers.get("База образования (9 или 11 классов)")],
                        'education_basis': row[headers.get("Основа образования (бюджет, внебюджет)")],
                        'admission_order': row[headers.get("Приказ о зачислении")],
                        'expell_order': row[headers.get("Приказ об отчислении")] if headers.get("Приказ об отчислении") is not None else None,
                        'transfer_to_2nd_year_order': row[headers.get("Переводной приказ на 2 курс")] if headers.get("Переводной приказ на 2 курс") is not None else None,
                        'transfer_to_3rd_year_order': row[headers.get("Переводной приказ на 3 курс")] if headers.get("Переводной приказ на 3 курс") is not None else None,
                        'transfer_to_4th_year_order': row[headers.get("Переводной приказ на 4 курс")] if headers.get("Переводной приказ на 4 курс") is not None else None,
                    }

                    if headers.get("Дата рождения") is not None and row[headers.get("Дата рождения")]:
                        cell_value = row[headers.get("Дата рождения")]
                        if isinstance(cell_value, datetime):
                            data['birth_date'] = cell_value.date()
                        else:
                            try:
                                data['birth_date'] = datetime.strptime(cell_value, "%d.%m.%Y").date()
                            except ValueError:
                                print(f"Ошибка даты у {data['full_name']}: {cell_value}")

                    # Проверка группы
                    group_filter = GroupStudents.objects.filter(full_name=data['group_number'])
                    if group_filter.exists():
                        data['group'] = group_filter.first()
                    else:
                        raise ValidationError(f'Группы {data["group_number"]} не существует')

                    # print(f"Добавляем: {data}")
                    self.data_list.append(data)
                # elif sheet_name == required_sheets[7]:
                #     pass
                # elif sheet_name == required_sheets[8]:
                #     pass
                # elif sheet_name == required_sheets[9]:
                #     pass

    def save(self):
        # print(f"Всего данных для сохранения: {len(self.data_list)}")
        for data in self.data_list:
            student, created = Student.objects.update_or_create(
                full_name=data['full_name'],
                defaults={
                    'birth_date': data['birth_date'],
                    'group': data['group'],
                    'education_basis': data['education_basis'],
                    'education_base': data['education_base'],
                    'admission_order': data['admission_order'],
                    'expell_order': data['expell_order'],
                    'transfer_to_2nd_year_order': data['transfer_to_2nd_year_order'],
                    'transfer_to_3rd_year_order': data['transfer_to_3rd_year_order'],
                    'transfer_to_4th_year_order': data['transfer_to_4th_year_order'],
                }
            )
            # print(f"{'Создан' if created else 'Обновлен'} студент: {student.full_name}")


class PromoteStudentsForm(forms.ModelForm):
    EDUCATION_BASE_CHOICES = (
        ("", "Выберите базу образования"),  # Пустой вариант
        ("основное общее", "Основное общее"),
        ("среднее общее", "Среднее общее"),
    )

    EDUCATION_BASIS_CHOICES = (
        ("", "Выберите основу обучения"),  # Пустой вариант
        ("бюджет", 'Бюджет'),
        ("внебюджет", "Внебюджет")
    )

    ACADEMIC_DEBTS = (
        ("", "Выберите наличие задолженностей"),  # Пустой вариант
        (True, 'Есть задолженности'),
        (False, 'Нет задолженностей')
    )

    CURRENT_COURSE = (
        ("", "Выберите курс"),  # Пустой вариант
        (1, '1'),
        (2, '2'),
        (3, '3'),
        (4, '4')
    )

    education_base = forms.ChoiceField(
        label='База образования',
        choices=EDUCATION_BASE_CHOICES,
        required=False,
    )
    education_basis = forms.ChoiceField(
        label='Основа обучения',
        choices=EDUCATION_BASIS_CHOICES,
        required=False
    )
    academic_debts = forms.ChoiceField(
        label='Наличие академических задолженностей',
        choices=ACADEMIC_DEBTS,
        required=False
    )
    current_course = forms.ChoiceField(
        label='Текущий курс',
        choices=CURRENT_COURSE,
        required=False
    )

    students = forms.ModelMultipleChoiceField(
        queryset=Student.objects.none(),
        label='Студенты',
        widget=forms.CheckboxSelectMultiple,
        required=False
    )

    order_number = forms.CharField(
        label='Номер приказа',
        widget=forms.TextInput(
            attrs={
                'placeholder': 'Номер переводного приказа'
            }
        )
    )

    class Meta:
        model = CurrentStudent
        fields = []

    def __init__(self, *args, **kwargs):
        properties = kwargs.pop('properties', {})
        super().__init__(*args, **kwargs)
        self._configure_new_instance(properties)

    def _configure_new_instance(self, properties):
        """Настройка для фильтрации студентов"""

        # Получаем параметры фильтрации из properties
        education_base = properties.get('education_base')
        education_basis = properties.get('education_basis')
        academic_debt = properties.get('academic_debts')
        current_course = properties.get('current_course')

        # Базовый пустой QuerySet
        students_qs = Student.objects.none()

        # Проверяем, указаны ли какие-либо параметры
        if education_base or education_basis or academic_debt or current_course:
            # Строим QuerySet постепенно
            students_qs = Student.objects.all()

            if education_base:
                students_qs = students_qs.filter(education_base=education_base)

            if education_basis:
                students_qs = students_qs.filter(education_basis=education_basis)

            if academic_debt != "":
                has_debt = True if academic_debt == 'True' else False
                students_qs = students_qs.filter(academic_debts=has_debt)

            if current_course:
                # students_qs = students_qs.filter(group__current_course=int(current_course))

                if int(current_course) == 1:
                    students_qs = students_qs.filter(transfer_to_2nd_year_order=None, transfer_to_3rd_year_order=None,
                                                     transfer_to_4th_year_order=None)
                if int(current_course) == 2:
                    students_qs = students_qs.filter(transfer_to_2nd_year_order__isnull=False, transfer_to_3rd_year_order=None,
                                                     transfer_to_4th_year_order=None)
                if int(current_course) == 3:
                    students_qs = students_qs.filter(transfer_to_2nd_year_order__isnull=False, transfer_to_3rd_year_order__isnull=False,
                                                     transfer_to_4th_year_order=None)

        # Устанавливаем отфильтрованный (или пустой) список студентов
        self.fields['students'].queryset = students_qs.order_by('full_name')





